from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup as bs
import openpyxl
# 정보입력란 - 추후 데이터베이스에 삽입하여, 관리자가 입력/수정하도록 변경
# 크롤링할 메일 정보
usr = ""    #계정 ID
mail_PW = "" #계정 PW
# 주소
url = "https://mw.genie.co.kr/login/login?rtnURL=https://mw.genie.co.kr/"
# 크롬드라이버 위치
path = 'chromedriver.exe'
# 로그인페이지 이동버튼
# login_btn = '//*[@id="btnLogin"]'
# 로그인창 아이디/비밀번호칸
login_idinput = '//*[@id="gnb_uxd"]'
login_pwinput = '//*[@id="gnb_uxx"]'
# 로그인창 로그인버튼 XPath
login_btnxp = '//*[@id="f_login"]/fieldset/div[5]/input'

url2 = 'https://www.genie.co.kr/'   # 로그인 후 이동할 페이지

# 마이뮤직 버튼
btn1 = '//*[@id="gnb"]/div/div[2]/button'
# 플레이리스트 버튼
btn2 = '//*[@id="myAlbumDetail"]/ul/li[4]/a'



# 크롬드라이버 세팅
# options = webdriver.ChromeOptions()
# options.add_argument('headless')
# options.add_argument('window-size=1920x1080')
# options.add_argument("disable-gpu")
# 크롬드라이버 실행
# driver = webdriver.Chrome(path, chrome_options=options)
driver = webdriver.Chrome(path)
driver.get(url)

# 로그인
# elem = driver.find_element_by_id(login_idinput)
elem = driver.find_element_by_xpath(login_idinput)
elem.send_keys(usr)
driver.implicitly_wait(0.5)
# elem = driver.find_element_by_id(login_pwinput)
elem = driver.find_element_by_xpath(login_pwinput)
elem.send_keys(mail_PW)
driver.implicitly_wait(0.5)
elem = driver.find_element_by_xpath(login_btnxp)
elem.click()
driver.implicitly_wait(3)

# pc용 사이트로 리디렉션
driver.get(url2)
# 플레이리스트
elem = driver.find_element_by_xpath(btn1)
elem.click()
driver.implicitly_wait(0.5)
elem = driver.find_element_by_xpath(btn2)
elem.click()
driver.implicitly_wait(3)

# 엑셀만들기
wb = openpyxl.Workbook()
st = wb.active


for n in [1, 2]:
    soup = bs(driver.page_source, 'html.parser')
    title = soup.select('table.list-wrap > tbody > tr.list > td.info > a.title')    # 제목
    artist = soup.select('table.list-wrap > tbody > tr.list > td.info > a.artist')    # 아티스트
    albumtitle = soup.select('table.list-wrap > tbody > tr.list > td.info > a.albumtitle')    # 앨범타이틀
    ind = 1
    for nn in title:
        st.cell(row=ind + (500*(n-1)), column=1).value = nn.text
        ind += 1

    ind = 1
    for nn in artist:
        st.cell(row=ind + (500*(n-1)), column=2).value = nn.text
        ind += 1
    
    ind = 1
    for nn in albumtitle:
        st.cell(row=ind + (500*(n-1)), column=3).value = nn.text
        ind += 1

    if n == 1:
        driver.get('https://www.genie.co.kr/myMusic/myfolder?mxnm=24665569&pg=2&pgsize=500')

wb.save('test.xlsx')
driver.close()
